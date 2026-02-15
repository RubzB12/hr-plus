"""Services for analytics app."""

import csv
import io
from datetime import datetime
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.core.exceptions import BusinessValidationError

from .selectors import ReportSelector


class ReportService:
    """Service for generating custom reports."""

    SUPPORTED_FORMATS = ['csv', 'excel', 'json']

    REPORT_TYPES = {
        'pipeline_conversion': ReportSelector.get_pipeline_conversion_report,
        'time_to_fill': ReportSelector.get_time_to_fill_report,
        'source_effectiveness': ReportSelector.get_source_effectiveness_report,
        'offer_analysis': ReportSelector.get_offer_analysis_report,
        'interviewer_calibration': ReportSelector.get_interviewer_calibration_report,
        'requisition_aging': ReportSelector.get_requisition_aging_report,
    }

    @staticmethod
    def generate_report(
        *,
        report_type: str,
        format: str = 'json',
        start_date=None,
        end_date=None,
        department_id=None,
        filters: dict = None,
    ):
        """
        Generate a report of the specified type.

        Args:
            report_type: Type of report to generate
            format: Output format (csv, excel, json)
            start_date: Start date for report data
            end_date: End date for report data
            department_id: Filter by department
            filters: Additional filters

        Returns:
            Report data in specified format

        Raises:
            BusinessValidationError: If report type or format is invalid
        """
        if report_type not in ReportService.REPORT_TYPES:
            raise BusinessValidationError(
                f'Invalid report type. Must be one of: {", ".join(ReportService.REPORT_TYPES.keys())}',
            )

        if format not in ReportService.SUPPORTED_FORMATS:
            raise BusinessValidationError(
                f'Invalid format. Must be one of: {", ".join(ReportService.SUPPORTED_FORMATS)}',
            )

        # Get report generator function
        generator = ReportService.REPORT_TYPES[report_type]

        # Build kwargs for generator
        kwargs = {}
        if report_type != 'requisition_aging':
            # Most reports need date range
            if not start_date:
                start_date = timezone.now() - timezone.timedelta(days=30)
            if not end_date:
                end_date = timezone.now()

            kwargs['start_date'] = start_date
            kwargs['end_date'] = end_date

        if department_id and report_type not in ['interviewer_calibration', 'requisition_aging']:
            kwargs['department_id'] = department_id

        # Generate report data
        data = generator(**kwargs)

        # Format output
        if format == 'json':
            return data
        elif format == 'csv':
            return ReportService._format_as_csv(report_type, data)
        elif format == 'excel':
            return ReportService._format_as_excel(report_type, data)

    @staticmethod
    def _format_as_csv(report_type: str, data: dict) -> str:
        """Format report data as CSV."""
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type == 'pipeline_conversion':
            # Header
            writer.writerow(['Pipeline Conversion Report'])
            writer.writerow(['Period', f"{data['period']['start']} to {data['period']['end']}"])
            writer.writerow(['Total Applications', data['total_applications']])
            writer.writerow([])

            # Data
            writer.writerow(['Stage', 'Count', 'Conversion Rate (%)'])
            for conv in data['conversions']:
                writer.writerow([
                    conv['stage'],
                    conv['count'],
                    conv['conversion_rate'],
                ])

        elif report_type == 'time_to_fill':
            writer.writerow(['Time to Fill Report'])
            writer.writerow(['Period', f"{data['period']['start']} to {data['period']['end']}"])
            writer.writerow(['Overall Average Days', data['overall_avg_days']])
            writer.writerow([])

            writer.writerow(['By Department'])
            writer.writerow(['Department', 'Avg Days', 'Positions Filled'])
            for dept in data['by_department']:
                writer.writerow([dept['department'], dept['avg_days'], dept['positions_filled']])

            writer.writerow([])
            writer.writerow(['By Level'])
            writer.writerow(['Level', 'Avg Days', 'Positions Filled'])
            for level in data['by_level']:
                writer.writerow([level['level'], level['avg_days'], level['positions_filled']])

        elif report_type == 'source_effectiveness':
            writer.writerow(['Source Effectiveness Report'])
            writer.writerow(['Period', f"{data['period']['start']} to {data['period']['end']}"])
            writer.writerow([])

            writer.writerow([
                'Source',
                'Applications',
                'Screen Rate (%)',
                'Interview Rate (%)',
                'Offer Rate (%)',
                'Hire Rate (%)',
                'Hires',
            ])
            for source in data['sources']:
                writer.writerow([
                    source['source'],
                    source['applications'],
                    source['screen_rate'],
                    source['interview_rate'],
                    source['offer_rate'],
                    source['hire_rate'],
                    source['hires'],
                ])

        elif report_type == 'offer_analysis':
            writer.writerow(['Offer Analysis Report'])
            writer.writerow(['Period', f"{data['period']['start']} to {data['period']['end']}"])
            writer.writerow([])

            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Offers', data['total_offers']])
            writer.writerow(['Accepted', data['accepted']])
            writer.writerow(['Declined', data['declined']])
            writer.writerow(['Pending', data['pending']])
            writer.writerow(['Withdrawn', data['withdrawn']])
            writer.writerow(['Acceptance Rate (%)', data['acceptance_rate']])
            writer.writerow(['Decline Rate (%)', data['decline_rate']])

            if data['decline_reasons']:
                writer.writerow([])
                writer.writerow(['Decline Reasons'])
                writer.writerow(['Reason', 'Count'])
                for reason in data['decline_reasons']:
                    writer.writerow([reason['declined_reason'], reason['count']])

        elif report_type == 'interviewer_calibration':
            writer.writerow(['Interviewer Calibration Report'])
            writer.writerow(['Period', f"{data['period']['start']} to {data['period']['end']}"])
            writer.writerow([])

            writer.writerow(['Overall Stats'])
            writer.writerow(['Total Scorecards', data['total_scorecards']])
            writer.writerow(['Average Rating', data['avg_overall_rating']])
            writer.writerow([])

            writer.writerow(['Interviewer', 'Scorecards', 'Avg Rating', 'Hire Rate (%)'])
            for interviewer in data['interviewers']:
                writer.writerow([
                    interviewer['interviewer'],
                    interviewer['scorecards_count'],
                    interviewer['avg_rating'],
                    interviewer['hire_rate'],
                ])

        elif report_type == 'requisition_aging':
            writer.writerow(['Requisition Aging Report'])
            writer.writerow(['Total Open Requisitions', data['total_open']])
            writer.writerow([])

            writer.writerow([
                'Req ID',
                'Title',
                'Department',
                'Hiring Manager',
                'Opened At',
                'Days Open',
                'Applications',
            ])
            for req in data['all_requisitions']:
                writer.writerow([
                    req['requisition_id'],
                    req['title'],
                    req['department'],
                    req['hiring_manager'],
                    req['opened_at'],
                    req['days_open'],
                    req['applications'],
                ])

        return output.getvalue()

    @staticmethod
    def _format_as_excel(report_type: str, data: dict) -> bytes:
        """Format report data as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()

        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        row = 1

        if report_type == 'pipeline_conversion':
            # Title
            ws[f'A{row}'] = 'Pipeline Conversion Report'
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1

            # Period
            ws[f'A{row}'] = 'Period'
            ws[f'B{row}'] = f"{data['period']['start']} to {data['period']['end']}"
            row += 1

            ws[f'A{row}'] = 'Total Applications'
            ws[f'B{row}'] = data['total_applications']
            row += 2

            # Headers
            headers = ['Stage', 'Count', 'Conversion Rate (%)']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            row += 1

            # Data
            for conv in data['conversions']:
                ws.cell(row=row, column=1, value=conv['stage'])
                ws.cell(row=row, column=2, value=conv['count'])
                ws.cell(row=row, column=3, value=conv['conversion_rate'])
                row += 1

        elif report_type == 'source_effectiveness':
            ws[f'A{row}'] = 'Source Effectiveness Report'
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1

            ws[f'A{row}'] = 'Period'
            ws[f'B{row}'] = f"{data['period']['start']} to {data['period']['end']}"
            row += 2

            headers = [
                'Source',
                'Applications',
                'Screen Rate (%)',
                'Interview Rate (%)',
                'Offer Rate (%)',
                'Hire Rate (%)',
                'Hires',
            ]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            row += 1

            for source in data['sources']:
                ws.cell(row=row, column=1, value=source['source'])
                ws.cell(row=row, column=2, value=source['applications'])
                ws.cell(row=row, column=3, value=source['screen_rate'])
                ws.cell(row=row, column=4, value=source['interview_rate'])
                ws.cell(row=row, column=5, value=source['offer_rate'])
                ws.cell(row=row, column=6, value=source['hire_rate'])
                ws.cell(row=row, column=7, value=source['hires'])
                row += 1

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ReportScheduleService:
    """Service for scheduling recurring reports."""

    @staticmethod
    @transaction.atomic
    def schedule_report(
        *,
        report_type: str,
        format: str,
        frequency: str,
        recipients: list,
        department_id=None,
        created_by,
    ):
        """
        Schedule a recurring report.

        Args:
            report_type: Type of report
            format: Output format
            frequency: Frequency (daily, weekly, monthly)
            recipients: List of email addresses
            department_id: Filter by department
            created_by: User scheduling the report

        Returns:
            Scheduled report configuration (placeholder - would create Celery Beat schedule)
        """
        # Validate inputs
        if report_type not in ReportService.REPORT_TYPES:
            raise BusinessValidationError('Invalid report type')

        if format not in ReportService.SUPPORTED_FORMATS:
            raise BusinessValidationError('Invalid format')

        if frequency not in ['daily', 'weekly', 'monthly']:
            raise BusinessValidationError('Frequency must be daily, weekly, or monthly')

        if not recipients:
            raise BusinessValidationError('At least one recipient is required')

        # In a real implementation, this would:
        # 1. Create a PeriodicTask in django-celery-beat
        # 2. Set up the schedule based on frequency
        # 3. Store configuration in a ReportSchedule model

        # For now, return configuration
        return {
            'report_type': report_type,
            'format': format,
            'frequency': frequency,
            'recipients': recipients,
            'department_id': department_id,
            'created_by': created_by.id,
            'created_at': timezone.now().isoformat(),
            'status': 'active',
        }
